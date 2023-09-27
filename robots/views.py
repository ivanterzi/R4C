import json
import re
from datetime import datetime, timedelta, timezone
from django.views.decorators.csrf import csrf_exempt
from django.shortcuts import render
from .models import Robot
from django.http import JsonResponse, HttpResponse
import openpyxl
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
from urllib.parse import quote
from collections import defaultdict

# Регулярное выражение для проверки формата модели (два символа, буквы и цифры)
MODEL_REGEX = r'^[A-Z0-9]{2}$'


@csrf_exempt
def create_robot(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            model = data.get('model')
            version = data.get('version')
            created_str = data.get('created')

            # Проверка формата модели с использованием регулярного выражения
            if not re.match(MODEL_REGEX, model):
                return JsonResponse({'error': 'Invalid model format'}, status=400)

            if version and created_str:
                try:
                    # Попытка преобразовать строку даты в объект datetime
                    created = datetime.strptime(created_str, '%Y-%m-%d %H:%M:%S')
                except ValueError:
                    return JsonResponse({'error': 'Invalid date format'}, status=400)

                robot = Robot(model=model, version=version, created=created)
                robot.save()
                return JsonResponse({'message': 'Robot record created successfully'}, status=201)
            else:
                return JsonResponse({'error': 'Invalid data'}, status=400)
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON data'}, status=400)


def get_robots(request):
    if request.method == 'GET':
        # Получите все уникальные модели роботов из базы данных
        models = Robot.objects.values_list('model', flat=True).distinct()

        # Создайте новую книгу Excel
        workbook = Workbook()

        # Для каждой модели создайте новый лист
        for model in models:
            worksheet = workbook.create_sheet(title=model)

            # Заголовок для листа
            worksheet.append(["Модель", "Версия", "Количество за неделю"])

            # Получите роботов для данной модели из базы данных
            robots = Robot.objects.filter(model=model)

            # Создайте словарь для подсчета количества роботов по версиям
            version_count = defaultdict(int)

            # Определите начало и конец периода (неделя)

            end_date = datetime.now(timezone.utc)
            start_date = end_date - timedelta(days=7)

            # Подсчитайте количество роботов за неделю
            for robot in robots:
                if start_date <= robot.created <= end_date:
                    version_count[robot.version] += 1

            # Заполните лист данными
            for version, count in version_count.items():
                worksheet.append([model, version, count])

        # Удалите стандартный лист
        del workbook[workbook.sheetnames[0]]
        filename = "сводка_по_роботам.xlsx"
        # Создайте HTTP-ответ для скачивания Excel-файла
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{quote(filename)}"'

        # Сохраните книгу Excel в HTTP-ответ
        workbook.save(response)

        return response